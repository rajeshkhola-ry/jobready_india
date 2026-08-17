#!/usr/bin/env python3
"""Exports every sheet of an .xlsx workbook to its own UTF-8 CSV file.

Usage: xlsx_to_csv.py <input.xlsx> <output_dir>

Uses openpyxl's read-only/streaming mode so row count is never capped -
suitable for 50,000+ row billing/GSTR sheets without loading the whole
workbook into memory. Prints a JSON summary to stdout:
    {"sheets": ["<output_dir>/Sheet1.csv", ...]}
or, on failure:
    {"error": "<message>"}
"""

import csv
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError as exc:
    print(json.dumps({'error': 'openpyxl not installed: ' + str(exc)}))
    sys.exit(1)


def sanitize_sheet_filename(name):
    safe = ''.join(ch if ch.isalnum() or ch in (' ', '-', '_') else '_' for ch in name).strip()
    return safe or 'Sheet'


def main():
    if len(sys.argv) < 3:
        print(json.dumps({'error': 'Usage: xlsx_to_csv.py <input.xlsx> <output_dir>'}))
        sys.exit(1)

    input_path = sys.argv[1]
    output_dir = sys.argv[2]

    try:
        os.makedirs(output_dir, exist_ok=True)
        workbook = load_workbook(filename=input_path, read_only=True, data_only=True)
    except Exception as exc:
        print(json.dumps({'error': 'Cannot open workbook: ' + str(exc)}))
        sys.exit(1)

    written_files = []

    try:
        for index, sheet_name in enumerate(workbook.sheetnames):
            sheet = workbook[sheet_name]
            safe_name = sanitize_sheet_filename(sheet_name) or ('Sheet' + str(index + 1))

            candidate = os.path.join(output_dir, safe_name + '.csv')
            suffix = 1
            while candidate in written_files:
                suffix += 1
                candidate = os.path.join(output_dir, safe_name + '_' + str(suffix) + '.csv')

            with open(candidate, 'w', newline='', encoding='utf-8-sig') as handle:
                writer = csv.writer(handle)
                for row in sheet.iter_rows(values_only=True):
                    writer.writerow(['' if value is None else value for value in row])

            written_files.append(candidate)
    except Exception as exc:
        print(json.dumps({'error': 'Failed while exporting sheets: ' + str(exc)}))
        sys.exit(1)
    finally:
        workbook.close()

    print(json.dumps({'sheets': written_files}))


if __name__ == '__main__':
    main()
